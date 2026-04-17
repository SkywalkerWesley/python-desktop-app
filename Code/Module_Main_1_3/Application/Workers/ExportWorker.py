from PyQt5 import QtCore
import openpyxl
from itertools import zip_longest

class ExportWorker(QtCore.QObject):
    """
    an object that export the sample plot table to and excle file
    """
    finished = QtCore.pyqtSignal(str)
    error = QtCore.pyqtSignal(str)

    def __init__(self, path, samplePlotData):
        super().__init__()
        self.path = path
        self.samplePlotData = samplePlotData

    @QtCore.pyqtSlot()
    def run(self):
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for d in self.samplePlotData:
                sheetName = str(d[0])[:31] if d[0] else "Sheet"
                ws = wb.create_sheet(title=sheetName)

                columnHeaders = [str(d[0])]
                dataLists = [[]]
                for i in range(1, len(d)):
                    header_name = str(d[i][0])
                    columnHeaders.append(header_name)
                    dataLists.append(d[i][1])

                bold_font = openpyxl.styles.Font(bold=True)
                for col, header in enumerate(columnHeaders, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = bold_font

                for rowIdx, rowValues in enumerate(zip_longest(*dataLists, fillvalue=''), start=2):
                    for colIdx, value in enumerate(rowValues, start=1):
                        if value is None:
                            value = ""
                        elif hasattr(value, "item"):
                            value = value.item()
                        elif not isinstance(value, (int, float, str)):
                            value = str(value)
                        ws.cell(row=rowIdx, column=colIdx, value=value)

                for colIdx, header in enumerate(columnHeaders, start=1):
                    colLetter = openpyxl.utils.get_column_letter(colIdx)
                    maxlength = len(header)
                    for v in dataLists[colIdx - 1]:
                        vStr = str(v)
                        if len(vStr) > maxlength:
                            maxlength = len(vStr)
                    ws.column_dimensions[colLetter].width = min(maxlength + 2, 50)

            wb.save(self.path)
            self.finished.emit(self.path)
        except Exception as e:
            self.error.emit(str(e))